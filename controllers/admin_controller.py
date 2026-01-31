# --- Importations nécessaires ---
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from database import (
    insert_schedule_slot,
    check_conflict,
    getConnection,
    DAYS
)


# Contrôleur pour l'administrateur

class AdminController:

    def __init__(self, admin_id):
        self.admin_id = admin_id

    def creer_creneau(self, course_id, instructor_id, group_id, room_id, day, start_hour, duration):
        conflit = check_conflict(instructor_id, group_id, room_id, day, start_hour, duration)
        if conflit:
            print(f" Conflit détecté : {conflit}")
            return False
        success = insert_schedule_slot(course_id, instructor_id, group_id, room_id, day, start_hour, duration, self.admin_id)
        if success:
            print(" Créneau ajouté avec succès.")
        return success

    def valider_reservation(self, reservation_id):
        conn = getConnection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE reservations
            SET status = 'APPROVED', approved_by = ?, approved_at = CURRENT_TIMESTAMP
            WHERE id = ? AND status = 'PENDING'
        """, (self.admin_id, reservation_id))
        conn.commit()
        if cursor.rowcount > 0:
            print(f" Réservation {reservation_id} validée.")
        else:
            print(f" Aucune réservation en attente avec l’ID {reservation_id}.")
        conn.close()

    def rejeter_reservation(self, reservation_id):
        conn = getConnection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE reservations
            SET status = 'REJECTED', approved_by = ?, approved_at = CURRENT_TIMESTAMP
            WHERE id = ? AND status = 'PENDING'
        """, (self.admin_id, reservation_id))
        conn.commit()
        if cursor.rowcount > 0:
            print(f" Réservation {reservation_id} rejetée.")
        else:
            print(f" Aucune réservation en attente avec l’ID {reservation_id}.")
        conn.close()

    def afficher_details_reservation(self, reservation_id):
        conn = getConnection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT r.id, u.full_name AS enseignant, rm.name AS salle, g.name AS groupe,
                   r.day, r.start_hour, r.duration, r.reason, r.status
            FROM reservations r
            JOIN instructors i ON r.instructor_id = i.id
            JOIN users u ON i.user_id = u.id
            LEFT JOIN rooms rm ON r.room_id = rm.id
            LEFT JOIN groups g ON r.group_id = g.id
            WHERE r.id = ?
        """, (reservation_id,))
        res = cursor.fetchone()
        conn.close()

        if res:
            print("\n Détails de la réservation :")
            print(f"- ID : {res['id']}")
            print(f"- Enseignant : {res['enseignant']}")
            print(f"- Salle : {res['salle']}")
            print(f"- Groupe : {res['groupe']}")
            print(f"- Jour : {DAYS[res['day']]}")
            print(f"- Heure début : {res['start_hour']}h")
            print(f"- Durée : {res['duration']}h")
            print(f"- Raison : {res['reason']}")
            print(f"- Statut : {res['status']}")
        else:
            print(" Réservation introuvable.")

    def afficher_reservations_en_attente(self):
        conn = getConnection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT r.id, u.full_name AS enseignant, r.day, r.start_hour, r.duration, r.reason
            FROM reservations r
            JOIN instructors i ON r.instructor_id = i.id
            JOIN users u ON i.user_id = u.id
            WHERE r.status = 'PENDING'
            ORDER BY r.day, r.start_hour
        """)
        results = cursor.fetchall()
        conn.close()

        if results:
            print("\n Réservations en attente :")
            for r in results:
                print(f"- ID {r['id']} | {DAYS[r['day']]} {r['start_hour']}h-{r['start_hour']+r['duration']}h | Enseignant : {r['enseignant']} | Raison : {r['reason']}")
        else:
            print("Aucune réservation en attente.")

    def afficher_statistiques(self):
        conn = getConnection()
        cursor = conn.cursor()

        nb_creneaux = cursor.execute("SELECT COUNT(*) FROM timetable").fetchone()[0]
        nb_reservations = cursor.execute("SELECT COUNT(*) FROM reservations WHERE status = 'APPROVED'").fetchone()[0]
        nb_salles = cursor.execute("SELECT COUNT(*) FROM rooms").fetchone()[0]

        print("\n Statistiques générales :")
        print(f"- Nombre total de créneaux planifiés : {nb_creneaux}")
        print(f"- Réservations approuvées : {nb_reservations}")
        print(f"- Nombre de salles disponibles : {nb_salles}")

        print("\n Taux d’occupation des salles :")
        cursor.execute("""
            SELECT r.name, COUNT(t.id) AS nb_creneaux
            FROM rooms r
            LEFT JOIN timetable t ON r.id = t.room_id
            GROUP BY r.id
            ORDER BY nb_creneaux DESC
        """)
        for row in cursor.fetchall():
            print(f"- {row['name']} : {row['nb_creneaux']} créneaux")

        conn.close()

    def exporter_statistiques_excel(self, filename="statistiques.xlsx"):
        conn = getConnection()
        cursor = conn.cursor()

        stats = cursor.execute("""
            SELECT r.name, COUNT(t.id) AS nb_creneaux
            FROM rooms r
            LEFT JOIN timetable t ON r.id = t.room_id
            GROUP BY r.id
        """).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statistiques"

        ws.append(["Salle", "Nombre de créneaux"])
        for row in stats:
            ws.append([row["name"], row["nb_creneaux"]])

        wb.save(filename)
        conn.close()
        print(f" Statistiques exportées vers {filename}")

    def exporter_statistiques_pdf(self, filename="statistiques.pdf"):
        conn = getConnection()
        cursor = conn.cursor()

        stats = cursor.execute("""
            SELECT r.name, COUNT(t.id) AS nb_creneaux
            FROM rooms r
            LEFT JOIN timetable t ON r.id = t.room_id
            GROUP BY r.id
        """).fetchall()

        c = canvas.Canvas(filename, pagesize=letter)
        c.drawString(100, 750, "Statistiques d'occupation des salles")

        y = 700
        for row in stats:
            c.drawString(100, y, f"{row['name']} : {row['nb_creneaux']} créneaux")
            y -= 20

        c.save()
        conn.close()
        print(f" Statistiques exportées vers {filename}")

    #Method inside the class (4 spaces indentation) ---
    def generer_planning_complet(self):
        """
        Génère l'emploi du temps complet en utilisant l'algorithme génétique.
        Cette action efface le planning existant pour une régénération propre.
        """
        print("Démarrage de la génération automatique...")
        
        # 1. Nettoyer la table timetable (Optionnel: on pourrait vouloir garder les trucs manuels)
        conn = getConnection()
        cursor = conn.cursor()
        cursor = conn.cursor()
        # cursor.execute("DELETE FROM timetable") # Reset complet pour la démo
        # conn.commit()
        conn.close()
        
        # 2. Lancer l'algo
        from Schedule import GeneticAlgorithm, Configuration, DAY_HOURS
        
        # Recharger la config pour être sûr d'avoir les dernières données
        config = Configuration.get_instance()
        config.load_data()
        
        if config.GetNumberOfCourseClasses() == 0:
            return "Aucun cours à planifier (Tables vides ?)"
            
        ga = GeneticAlgorithm(population_size=12, mutation_size=2)
        # On lance sur 50 générations (peut être ajusté)
        best_schedule = ga.evolve(max_generations=50, target_fitness=0.95)
        
        # 3. Sauvegarder le meilleur résultat
        conn = getConnection()
        cursor = conn.cursor()
        
        count = 0
        nr = config.GetNumberOfRooms()
        day_size = DAY_HOURS * nr
        
        for cc, pos in best_schedule.classes.items():
            # Décodage de la position
            day = pos // day_size
            rem = pos % day_size
            room_idx = rem // DAY_HOURS
            time = rem % DAY_HOURS
            
            # Récupération des IDs réels
            # Note: day est 0-indexed dans l'algo, mais 1-indexed dans la DB (Lundi=1)
            db_day = day + 1 
            
            # Ajustement de l'heure (L'algo commence à 0 -> 8h00)
            db_start_hour = 8 + time
            
            # Récupération de l'objet Salle réel
            room_wrapper = config.GetRoomById(room_idx)
            room_id = room_wrapper.GetId()
            
            # Données du cours
            subj = cc.GetSubject()
            grp = cc.GetGroups()[0] # On a simplifié à 1 groupe
            instr = cc.GetProfessor()
            duration = cc.GetDuration()
            
            # Insertion directe (on bypass check_conflict car l'algo l'a fait)
            try:
                cursor.execute("""
                    INSERT INTO timetable (course_id, instructor_id, group_id, room_id, day, start_hour, duration, created_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (subj['id'], instr['id'], grp['id'], room_id, db_day, db_start_hour, duration, self.admin_id))
                count += 1
            except Exception as e:
                print(f"Erreur insertion {subj['name']}: {e}")
                
        conn.commit()
        conn.close()
        
        return f"Génération terminée ! {count} cours planifiés avec un score de {best_schedule.fitness:.2%}."

    def affecter_automatiquement(self, subject_id, group_id, day, start_hour, duration):
        conn = getConnection()
        cursor = conn.cursor()

        # 1. Get Group info (student count)
        cursor.execute("SELECT student_count FROM groups WHERE id = ?", (group_id,))
        group = cursor.fetchone()
        if not group:
            conn.close()
            return "Group not found."

        # 2. Get Subject info (required equipment)
        cursor.execute("SELECT required_equipment FROM subjects WHERE id = ?", (subject_id,))
        subject = cursor.fetchone()
        
        # 3. Get all active rooms
        cursor.execute("SELECT * FROM rooms WHERE active = 1 ORDER BY capacity ASC")
        all_rooms = cursor.fetchall()

        # 4. Search for an available room
        assigned_room_id = None
        room_name = None
        for room in all_rooms:
            # Check capacity constraint
            if room['capacity'] < group['student_count']:
                continue
                
            # Check equipment constraint
            if subject['required_equipment']:
                if not room['equipments'] or subject['required_equipment'] not in room['equipments']:
                    continue

            # Check for schedule conflicts (Room availability)
            conflict = check_conflict(0, group_id, room['id'], day, start_hour, duration)
            
            if not conflict:
                assigned_room_id = room['id']
                room_name = room['name']
                break

        if assigned_room_id:
            # Find an available instructor
            cursor.execute("SELECT instructor_id FROM subject_instructors WHERE subject_id = ? LIMIT 1", (subject_id,))
            instr = cursor.fetchone()
            
            if instr:
                success = self.creer_creneau(subject_id, instr['instructor_id'], group_id, 
                                             assigned_room_id, day, start_hour, duration)
                conn.close()
                if success:
                    return f"Success: Room {room_name} assigned automatically."
            else:
                conn.close()
                return "Error: No qualified instructor found for this subject."
        
        conn.close()
        return "Error: No suitable room found for this slot."

    def exporter_planning_filiere_pdf(self, filiere_name, filename="Planning_FST.pdf"):
        """
        Génère un export PDF structuré selon le format officiel de l'Université Abdelmalek Essaâdi.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from datetime import datetime
            import os
        except ImportError:
            print("ReportLab n'est pas installé. Veuillez installer 'reportlab' via pip.")
            return "Erreur: ReportLab non installé"

        # Générer un nom de fichier unique avec horodatage
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = filename.replace('.pdf', '')
        unique_filename = f"{base_name}_{timestamp}.pdf"
        
        # Utiliser le dossier Documents si disponible
        try:
            documents_path = os.path.join(os.path.expanduser("~"), "Documents")
            if os.path.exists(documents_path):
                unique_filename = os.path.join(documents_path, unique_filename)
        except:
            pass  # Garder le nom local si erreur

        # Configuration du document en mode paysage
        doc = SimpleDocTemplate(unique_filename, pagesize=landscape(A4))
        elements = []
        styles = getSampleStyleSheet()

        # Styles pour l'en-tête institutionnel
        style_header = ParagraphStyle('Header', fontSize=12, leading=14, alignment=1)
        style_title = ParagraphStyle('Title', fontSize=13, leading=16, alignment=1, 
                                     spaceAfter=10, fontName='Helvetica-Bold')
        style_nb = ParagraphStyle('NB', fontSize=9, italic=True)

        # En-tête de l'université et de la faculté 
        elements.append(Paragraph(
            "Université Abdelmalek Essaâdi<br/>Faculté des Sciences et Techniques - Tanger", 
            style_header
        ))
        elements.append(Paragraph(f"Emploi du Temps du Semestre 6 (2025/2026)", style_title))
        elements.append(Paragraph(f"Filière : {filiere_name}", style_title))
        elements.append(Spacer(1, 15))

        # Créneaux horaires officiels de la FST Tanger
        time_slots = ["09h00-10h30", "10h45-12h15", "12h30-14h00", "14h15-15h45", "16h00-17h30"]
        data = [["JOURS"] + time_slots]
        days_list = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI", "SAMEDI"]

        # Mapping des jours pour la base de données
        DAYS_MAPPING = {
            "LUNDI": 1, "MARDI": 2, "MERCREDI": 3, 
            "JEUDI": 4, "VENDREDI": 5, "SAMEDI": 6
        }
        
        # Mapping des créneaux horaires vers les heures de la base
        SLOT_TO_HOUR = {
            "09h00-10h30": 9,
            "10h45-12h15": 10,
            "12h30-14h00": 12,
            "14h15-15h45": 14,
            "16h00-17h30": 16
        }

        conn = None
        try:
            conn = getConnection()
            cursor = conn.cursor()

            for day_name in days_list:
                row = [day_name]
                day_idx = DAYS_MAPPING.get(day_name, 0)
                
                for slot in time_slots:
                    # Gestion de l'exception du Vendredi après-midi (15h00)
                    if day_name == "VENDREDI" and slot == "14h15-15h45":
                        start_h = 15
                    else:
                        start_h = SLOT_TO_HOUR.get(slot, int(slot.split('h')[0]))
                    
                    # Récupération des cours, salles et groupes
                    cursor.execute("""
                        SELECT s.name as subject, r.name as room, g.name as group_name
                        FROM timetable t
                        JOIN subjects s ON t.course_id = s.id
                        JOIN rooms r ON t.room_id = r.id
                        JOIN groups g ON t.group_id = g.id
                        WHERE t.day = ? AND t.start_hour = ? AND g.filiere = ?
                        ORDER BY g.name
                    """, (day_idx, start_h, filiere_name))
                    
                    results = cursor.fetchall()
                    
                    if results:
                        cell_items = []
                        for res in results:
                            cell_items.append(f"{res['subject']} ({res['group_name']})\n{res['room']}")
                        row.append("\n\n".join(cell_items))
                    else:
                        row.append("")
                
                data.append(row)
            
        except Exception as e:
            error_style = ParagraphStyle('Error', textColor=colors.red, fontSize=10)
            elements.append(Paragraph(f"Erreur technique : {str(e)}", error_style))
        finally:
            if conn:
                conn.close()

        # Application du style au tableau
        if len(data) > 1:
            table = Table(data, colWidths=[80] + [135] * len(time_slots))
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 1), (0, -1), colors.whitesmoke),
            ]))
            elements.append(table)

        # Mentions obligatoires (N.B.)
        elements.append(Spacer(1, 15))
        nb_text = (
            "<b>N.B:</b> Les plannings de Travaux Pratiques seront affichés dans les départements concernés.<br/>"
            "Le Vendredi après-midi, les cours commencent à 15h00."
        )
        elements.append(Paragraph(nb_text, style_nb))

        # Pied de page avec horodatage
        gen_date = datetime.now().strftime("%d/%m/%Y %H:%M")
        elements.append(Spacer(1, 10))
        elements.append(Paragraph(f"Document généré le {gen_date}", 
                                ParagraphStyle('Footer', fontSize=7, alignment=2)))

        # Construction finale du PDF
        doc.build(elements)
        return f"PDF généré avec succès : {unique_filename}"

    def exporter_planning_filiere_excel(self, filiere_name, filename="Planning_FST.xlsx"):
        """
        Génère un export Excel de l'emploi du temps par filière.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from datetime import datetime
            import os
        except ImportError:
            return "Erreur: openpyxl non installé"

        # Générer un nom de fichier unique avec horodatage
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = filename.replace('.xlsx', '')
        unique_filename = f"{base_name}_{timestamp}.xlsx"
        
        # Utiliser le dossier Documents si disponible
        try:
            documents_path = os.path.join(os.path.expanduser("~"), "Documents")
            if os.path.exists(documents_path):
                unique_filename = os.path.join(documents_path, unique_filename)
        except:
            pass

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"EDT {filiere_name}"[:31]  # Excel limite à 31 caractères

        # Styles
        header_font = Font(bold=True, size=14)
        title_font = Font(bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        day_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # En-tête
        ws.merge_cells('A1:F1')
        ws['A1'] = "Université Abdelmalek Essaâdi - FST Tanger"
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align

        ws.merge_cells('A2:F2')
        ws['A2'] = f"Emploi du Temps - Filière: {filiere_name}"
        ws['A2'].font = title_font
        ws['A2'].alignment = center_align

        # Créneaux horaires
        time_slots = ["09h00-10h30", "10h45-12h15", "12h30-14h00", "14h15-15h45", "16h00-17h30"]
        days_list = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI", "SAMEDI"]
        DAYS_MAPPING = {"LUNDI": 1, "MARDI": 2, "MERCREDI": 3, "JEUDI": 4, "VENDREDI": 5, "SAMEDI": 6}
        SLOT_TO_HOUR = {"09h00-10h30": 9, "10h45-12h15": 10, "12h30-14h00": 12, "14h15-15h45": 14, "16h00-17h30": 16}

        # En-tête du tableau (ligne 4)
        ws.cell(row=4, column=1, value="JOURS").font = title_font
        ws.cell(row=4, column=1).fill = header_fill
        ws.cell(row=4, column=1).alignment = center_align
        ws.cell(row=4, column=1).border = thin_border

        for col, slot in enumerate(time_slots, start=2):
            cell = ws.cell(row=4, column=col, value=slot)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Données
        conn = getConnection()
        cursor = conn.cursor()

        for row_idx, day_name in enumerate(days_list, start=5):
            day_cell = ws.cell(row=row_idx, column=1, value=day_name)
            day_cell.font = title_font
            day_cell.fill = day_fill
            day_cell.alignment = center_align
            day_cell.border = thin_border

            day_idx = DAYS_MAPPING.get(day_name, 0)

            for col_idx, slot in enumerate(time_slots, start=2):
                if day_name == "VENDREDI" and slot == "14h15-15h45":
                    start_h = 15
                else:
                    start_h = SLOT_TO_HOUR.get(slot, int(slot.split('h')[0]))

                cursor.execute("""
                    SELECT s.name as subject, r.name as room, g.name as group_name
                    FROM timetable t
                    JOIN subjects s ON t.course_id = s.id
                    JOIN rooms r ON t.room_id = r.id
                    JOIN groups g ON t.group_id = g.id
                    WHERE t.day = ? AND t.start_hour = ? AND g.filiere = ?
                    ORDER BY g.name
                """, (day_idx, start_h, filiere_name))

                results = cursor.fetchall()
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align

                if results:
                    cell_text = "\n".join([f"{r['subject']} ({r['group_name']})\n{r['room']}" for r in results])
                    cell.value = cell_text
                else:
                    cell.value = ""

        conn.close()

        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 12
        for col in range(2, len(time_slots) + 2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

        # Ajuster les hauteurs de lignes
        for row in range(5, 5 + len(days_list)):
            ws.row_dimensions[row].height = 60

        # Pied de page
        footer_row = 5 + len(days_list) + 1
        ws.cell(row=footer_row, column=1, value=f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        wb.save(unique_filename)
        return f"Excel généré avec succès : {unique_filename}"

    def exporter_planning_filiere_image(self, filiere_name, filename="Planning_FST.png"):
        """
        Génère un export Image (PNG) de l'emploi du temps par filière.
        """
        try:
            from PIL import Image, ImageDraw, ImageFont
            from datetime import datetime
            import os
        except ImportError:
            return "Erreur: Pillow non installé. Installez-le avec: pip install Pillow"

        # Générer un nom de fichier unique avec horodatage
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = filename.replace('.png', '')
        unique_filename = f"{base_name}_{timestamp}.png"
        
        # Utiliser le dossier Documents si disponible
        try:
            documents_path = os.path.join(os.path.expanduser("~"), "Documents")
            if os.path.exists(documents_path):
                unique_filename = os.path.join(documents_path, unique_filename)
        except:
            pass

        # Configuration
        time_slots = ["09h00-10h30", "10h45-12h15", "12h30-14h00", "14h15-15h45", "16h00-17h30"]
        days_list = ["LUNDI", "MARDI", "MERCREDI", "JEUDI", "VENDREDI", "SAMEDI"]
        DAYS_MAPPING = {"LUNDI": 1, "MARDI": 2, "MERCREDI": 3, "JEUDI": 4, "VENDREDI": 5, "SAMEDI": 6}
        SLOT_TO_HOUR = {"09h00-10h30": 9, "10h45-12h15": 10, "12h30-14h00": 12, "14h15-15h45": 14, "16h00-17h30": 16}

        # Dimensions
        cell_width = 150
        cell_height = 80
        header_height = 100
        day_col_width = 100
        
        img_width = day_col_width + (cell_width * len(time_slots)) + 20
        img_height = header_height + (cell_height * len(days_list)) + 60

        # Créer l'image
        img = Image.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)

        # Polices (utiliser les polices par défaut)
        try:
            font_title = ImageFont.truetype("arial.ttf", 16)
            font_header = ImageFont.truetype("arial.ttf", 12)
            font_cell = ImageFont.truetype("arial.ttf", 9)
        except:
            font_title = ImageFont.load_default()
            font_header = ImageFont.load_default()
            font_cell = ImageFont.load_default()

        # Couleurs
        header_color = (68, 114, 196)
        day_color = (217, 226, 243)
        grid_color = (0, 0, 0)
        text_color = (0, 0, 0)

        # Titre
        draw.text((10, 10), "Université Abdelmalek Essaâdi - FST Tanger", fill=text_color, font=font_title)
        draw.text((10, 35), f"Emploi du Temps - Filière: {filiere_name}", fill=text_color, font=font_header)

        # Position de départ du tableau
        start_x = 10
        start_y = header_height - 20

        # En-tête des créneaux
        draw.rectangle([start_x, start_y, start_x + day_col_width, start_y + 30], fill=header_color)
        draw.text((start_x + 10, start_y + 8), "JOURS", fill='white', font=font_header)

        for i, slot in enumerate(time_slots):
            x = start_x + day_col_width + (i * cell_width)
            draw.rectangle([x, start_y, x + cell_width, start_y + 30], fill=header_color)
            draw.text((x + 10, start_y + 8), slot, fill='white', font=font_cell)

        # Données
        conn = getConnection()
        cursor = conn.cursor()

        for row_idx, day_name in enumerate(days_list):
            y = start_y + 30 + (row_idx * cell_height)
            
            # Colonne jour
            draw.rectangle([start_x, y, start_x + day_col_width, y + cell_height], fill=day_color, outline=grid_color)
            draw.text((start_x + 10, y + 30), day_name, fill=text_color, font=font_header)

            day_idx = DAYS_MAPPING.get(day_name, 0)

            for col_idx, slot in enumerate(time_slots):
                x = start_x + day_col_width + (col_idx * cell_width)
                
                if day_name == "VENDREDI" and slot == "14h15-15h45":
                    start_h = 15
                else:
                    start_h = SLOT_TO_HOUR.get(slot, int(slot.split('h')[0]))

                cursor.execute("""
                    SELECT s.name as subject, r.name as room, g.name as group_name
                    FROM timetable t
                    JOIN subjects s ON t.course_id = s.id
                    JOIN rooms r ON t.room_id = r.id
                    JOIN groups g ON t.group_id = g.id
                    WHERE t.day = ? AND t.start_hour = ? AND g.filiere = ?
                    ORDER BY g.name
                """, (day_idx, start_h, filiere_name))

                results = cursor.fetchall()
                
                # Dessiner la cellule
                draw.rectangle([x, y, x + cell_width, y + cell_height], outline=grid_color)

                if results:
                    text_y = y + 5
                    for r in results[:2]:  # Limiter à 2 entrées par cellule pour lisibilité
                        text = f"{r['subject'][:15]}"
                        draw.text((x + 5, text_y), text, fill=text_color, font=font_cell)
                        text_y += 12
                        draw.text((x + 5, text_y), f"({r['group_name']}) {r['room']}", fill=(100, 100, 100), font=font_cell)
                        text_y += 15

        conn.close()

        # Pied de page
        draw.text((10, img_height - 25), f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}", fill=(128, 128, 128), font=font_cell)

        img.save(unique_filename)
        return f"Image générée avec succès : {unique_filename}"
